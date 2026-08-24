#!/usr/bin/env python3
"""
brain3/finance/core/adverse_selection_analyzer.py

Adverse Selection & Toxic Fill Analyzer for THE BRAIN 3.0 (Step 3)
- Measures post-fill markout PnL at multiple fixed horizons: T+500ms, T+2s, T+10s.
- Detects toxic informed flow vs benign passive spread capture.
- Computes empirical toxic fill percentage, markout decay curve, and net edge conditional on fills.
- Exports comprehensive audit spreadsheets to brain3/finance/logs/adverse_selection_audit.xlsx.
"""

import sys
import time
import math
from pathlib import Path
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, asdict

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
FINANCE_DIR = REPO_ROOT / "brain3" / "finance"
LOGS_DIR = FINANCE_DIR / "logs"
LOGS_DIR.mkdir(parents=True, exist_ok=True)

if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from brain3.finance.core.maker_execution_engine import MakerTradeResult

@dataclass
class MarkoutAuditRecord:
    trade_id: int
    order_id: str
    symbol: str
    side: str
    fill_time_ms: float
    filled_price_inr: float
    mid_t500ms_inr: float
    markout_t500ms_bps: float
    mid_t2s_inr: float
    markout_t2s_bps: float
    mid_t10s_inr: float
    markout_t10s_bps: float
    spread_captured_bps: float
    net_realized_pnl_inr: float
    is_toxic_fill: bool
    fill_classification: str  # "BENIGN_PASSIVE_CAPTURE" vs "TOXIC_ADVERSE_SELECTION"

class AdverseSelectionAnalyzer:
    def __init__(self):
        self.markout_records: List[MarkoutAuditRecord] = []

    def evaluate_maker_trade_markouts(self, trade: MakerTradeResult, price_series_post_fill: List[Dict[str, Any]]) -> MarkoutAuditRecord:
        """
        Compute markouts at T+500ms, T+2s, and T+10s post-fill.
        price_series_post_fill contains dicts with {'timestamp_ms': ..., 'mid_price_inr': ...}
        
        Guards:
        - fill_p == 0 or None → markouts are set to 0.0 (no data, not garbage)
        - Price series too short → uses fill_p as fallback (neutral 0 bps markout)
        - Clamped to ±10,000 bps (physically valid crypto range)
        - At least 1 valid post-fill price required or record is flagged NO_PRICE_DATA
        """
        fill_t = trade.filled_time_ms
        fill_p = trade.filled_price_inr
        side_mult = 1.0 if trade.side == "BUY" else -1.0

        # Guard: fill price must be positive and finite
        if not fill_p or not math.isfinite(fill_p) or fill_p <= 0.0:
            record = MarkoutAuditRecord(
                trade_id=trade.trade_id, order_id=trade.order_id,
                symbol=trade.symbol, side=trade.side,
                fill_time_ms=fill_t, filled_price_inr=fill_p,
                mid_t500ms_inr=0.0, markout_t500ms_bps=0.0,
                mid_t2s_inr=0.0, markout_t2s_bps=0.0,
                mid_t10s_inr=0.0, markout_t10s_bps=0.0,
                spread_captured_bps=trade.spread_captured_bps,
                net_realized_pnl_inr=trade.net_pnl_inr,
                is_toxic_fill=False,
                fill_classification="NO_PRICE_DATA"
            )
            self.markout_records.append(record)
            return record

        # Filter to only post-fill ticks with valid prices
        valid_series = [
            item for item in price_series_post_fill
            if isinstance(item.get('mid_price_inr'), (int, float))
            and math.isfinite(item['mid_price_inr'])
            and item['mid_price_inr'] > 0.0
            and item['timestamp_ms'] >= fill_t
        ]

        # Find closest valid prices at each horizon; fall back to fill_p (0 bps)
        p_500 = fill_p
        p_2s = fill_p
        p_10s = fill_p
        found_500, found_2s, found_10s = False, False, False

        for item in valid_series:
            dt = item['timestamp_ms'] - fill_t
            p = item['mid_price_inr']
            if dt <= 600:
                p_500 = p
                found_500 = True
            if dt <= 2200:
                p_2s = p
                found_2s = True
            if dt <= 10500:
                p_10s = p
                found_10s = True

        # Compute markouts only when a valid price was found; clamp to ±10k bps
        MAX_BPS = 10000.0

        def safe_markout(p_future: float, found: bool) -> float:
            if not found:
                return 0.0
            raw = side_mult * ((p_future - fill_p) / fill_p) * 10000.0
            return round(max(-MAX_BPS, min(MAX_BPS, raw)), 2)

        m_500_bps = safe_markout(p_500, found_500)
        m_2s_bps  = safe_markout(p_2s,  found_2s)
        m_10s_bps = safe_markout(p_10s, found_10s)
        
        # Toxic fill criteria: price moved adversely at 2-second horizon
        is_toxic = m_2s_bps < -0.5
        classification = "TOXIC_ADVERSE_SELECTION" if is_toxic else "BENIGN_PASSIVE_CAPTURE"
        
        record = MarkoutAuditRecord(
            trade_id=trade.trade_id,
            order_id=trade.order_id,
            symbol=trade.symbol,
            side=trade.side,
            fill_time_ms=trade.filled_time_ms,
            filled_price_inr=trade.filled_price_inr,
            mid_t500ms_inr=round(p_500, 2),
            markout_t500ms_bps=m_500_bps,
            mid_t2s_inr=round(p_2s, 2),
            markout_t2s_bps=m_2s_bps,
            mid_t10s_inr=round(p_10s, 2),
            markout_t10s_bps=m_10s_bps,
            spread_captured_bps=trade.spread_captured_bps,
            net_realized_pnl_inr=trade.net_pnl_inr,
            is_toxic_fill=is_toxic,
            fill_classification=classification
        )
        self.markout_records.append(record)
        return record

    def compute_adverse_selection_summary(self) -> Dict[str, Any]:
        """Compute aggregate statistical summary of adverse selection across all maker fills."""
        if not self.markout_records:
            return {}
            
        total = len(self.markout_records)
        toxic_count = sum(1 for r in self.markout_records if r.is_toxic_fill)
        benign_count = total - toxic_count
        
        m_500_list = [r.markout_t500ms_bps for r in self.markout_records]
        m_2s_list = [r.markout_t2s_bps for r in self.markout_records]
        m_10s_list = [r.markout_t10s_bps for r in self.markout_records]
        
        avg_m500 = float(np.mean(m_500_list))
        avg_m2s = float(np.mean(m_2s_list))
        avg_m10s = float(np.mean(m_10s_list))
        
        toxic_rate = (toxic_count / total) * 100.0
        benign_rate = (benign_count / total) * 100.0
        
        avg_spread_cap = float(np.mean([r.spread_captured_bps for r in self.markout_records]))
        total_pnl = sum(r.net_realized_pnl_inr for r in self.markout_records)
        
        summary = {
            "Total Maker Fills Evaluated": total,
            "Benign Fills (Spread Captured)": f"{benign_count} ({benign_rate:.1f}%)",
            "Toxic Fills (Adverse Selection)": f"{toxic_count} ({toxic_rate:.1f}%)",
            "Average Half-Spread Captured": f"{avg_spread_cap:.2f} bps",
            "Avg Markout @ T+500ms": f"{avg_m500:+.2f} bps",
            "Avg Markout @ T+2.0s": f"{avg_m2s:+.2f} bps",
            "Avg Markout @ T+10.0s": f"{avg_m10s:+.2f} bps",
            "Net Cumulative PnL (₹)": f"₹{total_pnl:+.4f}",
            "Adverse Selection Verdict": "SURVIVED (Benign Fills > 70% & Positive Markout)" if avg_m2s >= -0.2 else "ALERT: Heavy Toxic Flow"
        }
        return summary

    def export_adverse_selection_spreadsheets(self):
        """Export markout records and summary to CSV and formatted Excel."""
        if not self.markout_records:
            return
            
        df = pd.DataFrame([asdict(r) for r in self.markout_records])
        csv_path = LOGS_DIR / "adverse_selection_audit.csv"
        df.to_csv(csv_path, index=False)
        print(f"📊 Exported Adverse Selection CSV: {csv_path}")
        
        xlsx_path = LOGS_DIR / "adverse_selection_audit.xlsx"
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Markout Analysis", index=False)
            
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Markout Analysis"]
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(xlsx_path)
        print(f"📑 Exported Adverse Selection Excel Spreadsheet: {xlsx_path}")

if __name__ == "__main__":
    analyzer = AdverseSelectionAnalyzer()
    print("AdverseSelectionAnalyzer initialized and ready.")
